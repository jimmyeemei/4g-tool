from copy import copy
from io import BytesIO
import re

import openpyxl
import streamlit as st
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter


PLAN_SHEET_NAME = "RANCM-sdrPlan"
RRU_SHEET_NAME = "RRU"
DATA_START_ROW_RANCM = 6
times_font = Font(name="Times New Roman")

MODE_ALIASES = ("制式", "FDD/TDD")
EXPANSION_TYPE_ALIASES = ("扩容类型",)

CROSS_MODE_KEYS = {"tddtofdd", "fddtotdd"}
TDD_TARGET_MODE_KEYS = {"tddtotdd", "fddtotdd"}
HARD_EXPANSION_KEYS = {"硬扩"}
SOFT_EXPANSION_KEYS = {"软扩"}


def normalize_key(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"[\s_\-:/\\()\[\]{}（）【】]+", "", text)
    return text.lower()


def get_sheet_by_name_fuzzy(workbook, target_name):
    if target_name in workbook.sheetnames:
        return workbook[target_name]

    target_key = normalize_key(target_name)
    for sheet_name in workbook.sheetnames:
        if normalize_key(sheet_name) == target_key:
            return workbook[sheet_name]
    return None


def get_used_max_col(worksheet, header_rows=(1, 2, 3, 4, 5), max_scan=256):
    upper = min(worksheet.max_column, max_scan)
    used_max_col = 1
    for row_idx in header_rows:
        if row_idx > worksheet.max_row:
            continue
        for col_idx in range(1, upper + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value not in (None, ""):
                used_max_col = col_idx
    return used_max_col


def sanitize_punctuation(workbook):
    punct_map = str.maketrans(
        "，。：；！？（）【】“”‘’",
        ",.:;!?()[]\"\"''",
    )
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.translate(punct_map)


def set_row_font(worksheet, row_idx, font):
    for cell in worksheet[row_idx]:
        cell.font = font


def capture_row_snapshot(worksheet, row_idx, max_col=None):
    if max_col is None:
        max_col = get_used_max_col(worksheet)

    snapshot = []
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        snapshot.append(
            {
                "value": cell.value,
                "style": copy(cell._style),
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }
        )
    return {"cells": snapshot, "height": worksheet.row_dimensions[row_idx].height}


def write_row_snapshot(worksheet, row_idx, snapshot):
    if snapshot.get("height") is not None:
        worksheet.row_dimensions[row_idx].height = snapshot["height"]

    for col_idx, cell_data in enumerate(snapshot["cells"], start=1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        cell.value = cell_data["value"]
        cell._style = copy(cell_data["style"])
        cell.font = copy(cell_data["font"])
        cell.fill = copy(cell_data["fill"])
        cell.border = copy(cell_data["border"])
        cell.alignment = copy(cell_data["alignment"])
        cell.number_format = cell_data["number_format"]
        cell.protection = copy(cell_data["protection"])


def prepare_template_rows(worksheet, start_row, target_count, max_col=None):
    if max_col is None:
        max_col = get_used_max_col(worksheet)

    if target_count <= 0:
        if worksheet.max_row >= start_row:
            worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)
        return

    for col_idx in range(1, max_col + 1):
        last_valid_row = start_row - 1
        for row_idx in range(start_row, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value).strip() != "":
                last_valid_row = row_idx
            else:
                break

        if last_valid_row < start_row:
            last_valid_row = start_row

        current_count = last_valid_row - start_row + 1
        if current_count < target_count:
            source_cell = worksheet.cell(row=last_valid_row, column=col_idx)
            source_value = source_cell.value
            source_style = copy(source_cell._style)
            source_font = copy(source_cell.font)
            source_fill = copy(source_cell.fill)
            source_border = copy(source_cell.border)
            source_alignment = copy(source_cell.alignment)
            source_number_format = source_cell.number_format
            source_protection = copy(source_cell.protection)

            for row_idx in range(last_valid_row + 1, start_row + target_count):
                target_cell = worksheet.cell(row=row_idx, column=col_idx)
                target_cell.value = source_value
                target_cell._style = copy(source_style)
                target_cell.font = copy(source_font)
                target_cell.fill = copy(source_fill)
                target_cell.border = copy(source_border)
                target_cell.alignment = copy(source_alignment)
                target_cell.number_format = source_number_format
                target_cell.protection = copy(source_protection)

    valid_max_row = start_row + target_count - 1
    if worksheet.max_row > valid_max_row:
        worksheet.delete_rows(valid_max_row + 1, worksheet.max_row - valid_max_row)


def count_existing_rows(worksheet, start_row=DATA_START_ROW_RANCM, col_ref="A"):
    col_idx = column_index_from_string(col_ref) if isinstance(col_ref, str) else col_ref
    count = 0
    for row_idx in range(start_row, worksheet.max_row + 1):
        value = worksheet.cell(row=row_idx, column=col_idx).value
        if value is None or str(value).strip() == "":
            break
        count += 1
    return count


def get_cell_value(worksheet, row_idx, col_ref):
    col_idx = column_index_from_string(col_ref) if isinstance(col_ref, str) else col_ref
    return worksheet.cell(row=row_idx, column=col_idx).value


def set_cell_value(worksheet, row_idx, col_ref, value):
    col_idx = column_index_from_string(col_ref) if isinstance(col_ref, str) else col_ref
    worksheet.cell(row=row_idx, column=col_idx).value = value


def clear_row_columns(worksheet, row_idx, col_refs):
    for col_ref in col_refs:
        set_cell_value(worksheet, row_idx, col_ref, None)


def parse_int(value, default=0):
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"-?\d+", str(value))
    if match:
        return int(match.group())
    return default


def load_workbook_from_upload(uploaded_file, data_only=False):
    uploaded_file.seek(0)
    return openpyxl.load_workbook(uploaded_file, data_only=data_only)


def get_sheet_header_map(worksheet, header_row=1, max_scan=256):
    header_map = {}
    upper = min(worksheet.max_column, max_scan)
    for col_idx in range(1, upper + 1):
        value = worksheet.cell(row=header_row, column=col_idx).value
        if value is None or str(value).strip() == "":
            continue
        header_map[normalize_key(value)] = col_idx
    return header_map


def parse_plan_workbook(uploaded_file):
    workbook = load_workbook_from_upload(uploaded_file, data_only=True)
    worksheet = get_sheet_by_name_fuzzy(workbook, PLAN_SHEET_NAME)
    if worksheet is None:
        raise ValueError(f"未找到参数页: {PLAN_SHEET_NAME}")

    headers = {}
    header_map = {}
    for col_idx in range(1, worksheet.max_column + 1):
        header_value = worksheet.cell(row=1, column=col_idx).value
        if header_value is None or str(header_value).strip() == "":
            continue
        header_text = str(header_value).strip()
        headers[col_idx] = header_text
        header_map[normalize_key(header_text)] = header_text

    data_map = {header: [] for header in headers.values()}
    for col_idx, header_text in headers.items():
        values = []
        for row_idx in range(4, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            values.append("" if value is None else value)
        while values and values[-1] == "":
            values.pop()
        data_map[header_text] = values

    return {
        "workbook": workbook,
        "sheet": worksheet,
        "rru_sheet": get_sheet_by_name_fuzzy(workbook, RRU_SHEET_NAME),
        "data_map": data_map,
        "header_map": header_map,
    }


def normalize_aliases(aliases):
    if isinstance(aliases, str):
        return (aliases,)
    return tuple(aliases)


def get_plan_values(plan_data, aliases):
    aliases = normalize_aliases(aliases)
    for alias in aliases:
        real_name = plan_data["header_map"].get(normalize_key(alias))
        if real_name is not None:
            return list(plan_data["data_map"].get(real_name, [])), alias
    return [], aliases[0]


def get_plan_value(plan_data, aliases, index=0, repeat_single=True, default=""):
    values, _ = get_plan_values(plan_data, aliases)
    if not values:
        return default
    if len(values) == 1 and repeat_single:
        return values[0]
    if index < len(values):
        return values[index]
    return default


def get_plan_value_checked(
    plan_data,
    aliases,
    index,
    issues,
    sheet_name,
    column_name,
    repeat_single=True,
    required=True,
    default="",
):
    values, label = get_plan_values(plan_data, aliases)
    if not values:
        if required:
            issues.append(f"Sheet[{sheet_name}] 列[{column_name}] 参数[{label}]：源数据缺失。")
        return default

    if len(values) == 1 and repeat_single:
        value = values[0]
        row_desc = "第1行"
    elif index < len(values):
        value = values[index]
        row_desc = f"第{index + 1}行"
    else:
        if required:
            issues.append(
                f"Sheet[{sheet_name}] 列[{column_name}] 参数[{label}]：需要第{index + 1}行数据，但源数据只有{len(values)}行。"
            )
        return default

    if required and value in (None, ""):
        issues.append(f"Sheet[{sheet_name}] 列[{column_name}] 参数[{label}]：{row_desc}为空。")
        return default
    return value


def get_mode_value(plan_data):
    return str(get_plan_value(plan_data, MODE_ALIASES, 0, default="")).strip()


def get_mode_key(plan_data):
    return normalize_key(get_mode_value(plan_data))


def get_expansion_type_value(plan_data):
    return str(get_plan_value(plan_data, EXPANSION_TYPE_ALIASES, 0, default="")).strip()


def get_expansion_type_key(plan_data):
    return normalize_key(get_expansion_type_value(plan_data))


def is_cross_mode(plan_data):
    return get_mode_key(plan_data) in CROSS_MODE_KEYS


def is_hard_expansion(plan_data):
    return get_expansion_type_key(plan_data) in HARD_EXPANSION_KEYS


def is_soft_expansion(plan_data):
    return get_expansion_type_key(plan_data) in SOFT_EXPANSION_KEYS


def get_cell_count(plan_data, issues, context_name):
    value = get_plan_value_checked(plan_data, "cellnum", 0, issues, context_name, "cellnum")
    count = parse_int(value, 0)
    if count <= 0:
        issues.append(f"Sheet[{context_name}] 列[cellnum] 参数[cellnum]：未识别到有效的小区数。")
    return count


def get_ru_values(plan_data):
    values, _ = get_plan_values(plan_data, "RUDevice")
    return [value for value in values if value not in ("", None)]


def find_rru_row3_matches(rru_sheet, ru_device):
    matches = []
    if rru_sheet is None or ru_device in (None, ""):
        return matches

    used_max_col = get_used_max_col(rru_sheet, header_rows=(1, 2, 3), max_scan=256)
    target = str(ru_device).strip()
    for col_idx in range(2, used_max_col + 1):
        value = rru_sheet.cell(row=3, column=col_idx).value
        if value is not None and str(value).strip() == target:
            matches.append((3, col_idx))
    return matches


def find_rru_first_match(rru_sheet, ru_device):
    if rru_sheet is None or ru_device in (None, ""):
        return None

    used_max_col = get_used_max_col(rru_sheet, header_rows=(1, 2, 3), max_scan=256)
    target = str(ru_device).strip()
    for row_idx in range(3, rru_sheet.max_row + 1):
        for col_idx in range(2, used_max_col + 1):
            value = rru_sheet.cell(row=row_idx, column=col_idx).value
            if value is not None and str(value).strip() == target:
                return row_idx, col_idx
    return None


def collect_generated_ir_ant_group_numbers(rancm_bytes, issues):
    workbook = openpyxl.load_workbook(BytesIO(rancm_bytes), data_only=True)
    worksheet = get_sheet_by_name_fuzzy(workbook, "IrAntGroup")
    if worksheet is None:
        issues.append("RANCM结果中未找到Sheet[IrAntGroup]，无法生成cfgRadioNet的refIrAntGroup。")
        return []

    header_map = get_sheet_header_map(worksheet)
    col_idx = header_map.get(normalize_key("irAntGroupNo"))
    if col_idx is None:
        issues.append("RANCM结果Sheet[IrAntGroup]缺少表头[irAntGroupNo]。")
        return []

    values = []
    existing_rows = count_existing_rows(worksheet, DATA_START_ROW_RANCM, "A")
    for offset in range(existing_rows):
        row_idx = DATA_START_ROW_RANCM + offset
        value = worksheet.cell(row=row_idx, column=col_idx).value
        if value not in (None, ""):
            values.append(value)
    return values


def fill_managed_element(worksheet, plan_data, issues):
    if not is_cross_mode(plan_data):
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, 2)
    target_row = DATA_START_ROW_RANCM + 1
    clear_row_columns(worksheet, target_row, ["A", "F", "G", "H", "I", "AI"])

    set_cell_value(worksheet, target_row, "A", "M")
    for col_ref, header in [("F", "mimType"), ("G", "mimVersion"), ("H", "RADIOMODE"), ("I", "SWVERSION")]:
        value = get_plan_value_checked(plan_data, header, 0, issues, "ManagedElement", col_ref)
        set_cell_value(worksheet, target_row, col_ref, value)

    set_cell_value(worksheet, target_row, "AI", get_cell_value(worksheet, target_row, "H"))
    set_row_font(worksheet, target_row, times_font)


def fill_equipment(worksheet, plan_data, issues):
    if not is_hard_expansion(plan_data):
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, 1)
    row_idx = DATA_START_ROW_RANCM
    used_max_col = get_used_max_col(worksheet)
    clear_row_columns(worksheet, row_idx, [get_column_letter(col_idx) for col_idx in range(7, used_max_col + 1)])

    set_cell_value(worksheet, row_idx, "A", "M")
    for col_idx in range(7, used_max_col + 1):
        header_value = worksheet.cell(row=1, column=col_idx).value
        if header_value is None or str(header_value).strip() == "":
            continue
        value = get_plan_value(plan_data, str(header_value).strip(), 0, repeat_single=True, default="")
        if value not in (None, ""):
            worksheet.cell(row=row_idx, column=col_idx).value = value

    set_row_font(worksheet, row_idx, times_font)


def fill_ru(worksheet, plan_data, issues):
    if not is_hard_expansion(plan_data):
        return

    ru_values = get_ru_values(plan_data)
    target_count = len(ru_values)
    if target_count <= 0:
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, target_count)
    original_b_to_e = [get_cell_value(worksheet, DATA_START_ROW_RANCM, col_ref) for col_ref in ("B", "C", "D", "E")]

    for index in range(target_count):
        row_idx = DATA_START_ROW_RANCM + index
        clear_row_columns(worksheet, row_idx, ["A", "B", "C", "D", "E", "F", "G", "H", "K", "L", "M", "N"])
        set_cell_value(worksheet, row_idx, "A", "A")
        for col_ref, original_value in zip(("B", "C", "D", "E"), original_b_to_e):
            set_cell_value(worksheet, row_idx, col_ref, original_value)

        for col_ref, header in [
            ("F", "RUDevice"),
            ("G", "userLabel1"),
            ("H", "RUType"),
            ("K", "RADIOMODE"),
            ("L", "functionMode"),
            ("N", "connectModeWithUpRack"),
        ]:
            value = get_plan_value_checked(plan_data, header, index, issues, "RU", col_ref)
            set_cell_value(worksheet, row_idx, col_ref, value)

        set_cell_value(worksheet, row_idx, "M", get_cell_value(worksheet, row_idx, "F"))
        set_row_font(worksheet, row_idx, times_font)


def build_fiber_device_entries(plan_data, issues):
    entries = []
    rru_sheet = plan_data["rru_sheet"]
    for ru_device in get_ru_values(plan_data):
        matches = find_rru_row3_matches(rru_sheet, ru_device)
        if not matches:
            issues.append(f"Sheet[FiberDevice] 参数[RUDevice]：在模板Sheet[RRU]第3行未找到匹配值[{ru_device}]。")
            continue
        for _, col_idx in matches:
            entries.append({"ru_device": ru_device, "slot": rru_sheet.cell(row=1, column=col_idx).value, "col_idx": col_idx})
    return entries


def fill_fiber_device(worksheet, plan_data, issues):
    if not is_hard_expansion(plan_data):
        return

    entries = build_fiber_device_entries(plan_data, issues)
    target_count = len(entries)
    if target_count <= 0:
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, target_count)
    for index, entry in enumerate(entries):
        row_idx = DATA_START_ROW_RANCM + index
        clear_row_columns(worksheet, row_idx, ["A", "C", "D", "E", "F", "G", "H"])
        set_cell_value(worksheet, row_idx, "A", "M")
        for col_ref, header in [("C", "SubNetwork"), ("D", "ManagedElement"), ("E", "NE_Name")]:
            value = get_plan_value_checked(plan_data, header, 0, issues, "FiberDevice", col_ref)
            set_cell_value(worksheet, row_idx, col_ref, value)
        set_cell_value(worksheet, row_idx, "F", index + 1)
        set_cell_value(worksheet, row_idx, "G", f"1,1,{entry['slot']}")
        set_cell_value(worksheet, row_idx, "H", index)
        set_row_font(worksheet, row_idx, times_font)


def fill_fiber_cable(worksheet, plan_data, issues):
    if not is_hard_expansion(plan_data):
        return

    ru_values = get_ru_values(plan_data)
    target_count = len(ru_values)
    if target_count <= 0:
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, target_count)
    original_b_to_e = [get_cell_value(worksheet, DATA_START_ROW_RANCM, col_ref) for col_ref in ("B", "C", "D", "E")]

    for index, ru_device in enumerate(ru_values):
        row_idx = DATA_START_ROW_RANCM + index
        clear_row_columns(worksheet, row_idx, ["A", "B", "C", "D", "E", "F", "G", "H"])
        set_cell_value(worksheet, row_idx, "A", "A")
        for col_ref, original_value in zip(("B", "C", "D", "E"), original_b_to_e):
            set_cell_value(worksheet, row_idx, col_ref, original_value)

        ru_value = get_plan_value_checked(plan_data, "RUDevice", index, issues, "FiberCable", "F")
        set_cell_value(worksheet, row_idx, "F", ru_value)

        match = find_rru_first_match(plan_data["rru_sheet"], ru_value)
        if match is None:
            issues.append(f"Sheet[FiberCable] 列[G] 参数[ref1FiberDevice]：在模板Sheet[RRU]中未找到[{ru_value}]。")
        else:
            match_row, match_col = match
            if match_row == 3:
                x_value = plan_data["rru_sheet"].cell(row=1, column=match_col).value
                y_value = plan_data["rru_sheet"].cell(row=2, column=match_col).value
                set_cell_value(worksheet, row_idx, "G", f"(1,1,{x_value}):{y_value}")
            else:
                z_value = plan_data["rru_sheet"].cell(row=match_row - 1, column=match_col).value
                set_cell_value(worksheet, row_idx, "G", f"({z_value},1,1):2")

        set_cell_value(worksheet, row_idx, "H", f"({ru_value},1,1):1")
        set_row_font(worksheet, row_idx, times_font)


def fill_ir_ant_group(worksheet, plan_data, issues):
    if not is_hard_expansion(plan_data):
        return
    if get_mode_key(plan_data) not in TDD_TARGET_MODE_KEYS:
        return

    ru_values = get_ru_values(plan_data)
    target_count = len(ru_values)
    if target_count <= 0:
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, target_count)
    original_b_to_e = [get_cell_value(worksheet, DATA_START_ROW_RANCM, col_ref) for col_ref in ("B", "C", "D", "E")]
    start_ir_ant_group = parse_int(get_cell_value(worksheet, DATA_START_ROW_RANCM, "F"), 0)
    start_ant_entity = parse_int(get_cell_value(worksheet, DATA_START_ROW_RANCM, "I"), 0)

    for index in range(target_count):
        row_idx = DATA_START_ROW_RANCM + index
        clear_row_columns(worksheet, row_idx, ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"])
        set_cell_value(worksheet, row_idx, "A", "A")
        for col_ref, original_value in zip(("B", "C", "D", "E"), original_b_to_e):
            set_cell_value(worksheet, row_idx, col_ref, original_value)

        set_cell_value(worksheet, row_idx, "F", start_ir_ant_group + index + 1)
        set_cell_value(worksheet, row_idx, "G", 1)
        set_cell_value(worksheet, row_idx, "H", get_plan_value_checked(plan_data, "RUDevice", index, issues, "IrAntGroup", "H"))
        set_cell_value(worksheet, row_idx, "I", start_ant_entity + index + 1)
        set_cell_value(worksheet, row_idx, "J", get_plan_value_checked(plan_data, "refRfDevice", index, issues, "IrAntGroup", "J"))
        set_cell_value(worksheet, row_idx, "K", get_plan_value_checked(plan_data, "AntProfile", index, issues, "IrAntGroup", "K"))
        set_row_font(worksheet, row_idx, times_font)


def fill_ip_layer_config(worksheet, plan_data, issues):
    if not is_cross_mode(plan_data):
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, 2)
    base_snapshots = [capture_row_snapshot(worksheet, 6), capture_row_snapshot(worksheet, 7)]
    original_b_to_e = [get_cell_value(worksheet, DATA_START_ROW_RANCM, col_ref) for col_ref in ("B", "C", "D", "E")]

    for index in range(2):
        row_idx = DATA_START_ROW_RANCM + index
        write_row_snapshot(worksheet, row_idx, base_snapshots[index if index < len(base_snapshots) else 0])
        clear_row_columns(worksheet, row_idx, ["A", "B", "C", "D", "E", "J", "M", "N", "O"])
        set_cell_value(worksheet, row_idx, "A", "A")
        for col_ref, original_value in zip(("B", "C", "D", "E"), original_b_to_e):
            set_cell_value(worksheet, row_idx, col_ref, original_value)
        for col_ref, header in [("J", "vid"), ("M", "ipAddr"), ("N", "networkMask"), ("O", "gatewayIp")]:
            value = get_plan_value_checked(plan_data, header, index, issues, "IpLayerConfig", col_ref)
            set_cell_value(worksheet, row_idx, col_ref, value)
        set_row_font(worksheet, row_idx, times_font)


def fill_sctp(worksheet):
    target_count = 20
    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, target_count)
    base_snapshots = [capture_row_snapshot(worksheet, row_idx) for row_idx in range(6, 16)]
    start_sctp = parse_int(get_cell_value(worksheet, 15, "G"), 0)
    start_sctp_no = parse_int(get_cell_value(worksheet, 15, "H"), 0)

    for index, snapshot in enumerate(base_snapshots):
        row_idx = 16 + index
        write_row_snapshot(worksheet, row_idx, snapshot)
        set_cell_value(worksheet, row_idx, "G", start_sctp + index + 1)
        set_cell_value(worksheet, row_idx, "H", start_sctp_no + index + 1)
        set_row_font(worksheet, row_idx, times_font)


def fill_service_map(worksheet, plan_data, issues):
    if not is_cross_mode(plan_data):
        return

    prepare_template_rows(worksheet, DATA_START_ROW_RANCM, 4)
    base_snapshots = [capture_row_snapshot(worksheet, 6), capture_row_snapshot(worksheet, 7)]
    start_service_map_no = parse_int(get_cell_value(worksheet, 7, "F"), 0)

    for index in range(2):
        row_idx = 8 + index
        write_row_snapshot(worksheet, row_idx, base_snapshots[index if index < len(base_snapshots) else 0])
        clear_row_columns(worksheet, row_idx, ["F", "O", "P"])
        set_cell_value(worksheet, row_idx, "F", start_service_map_no + index + 1)
        set_cell_value(
            worksheet,
            row_idx,
            "O",
            get_plan_value_checked(plan_data, "fddServiceDscpMap", index, issues, "ServiceMap", "O"),
        )
        set_cell_value(
            worksheet,
            row_idx,
            "P",
            get_plan_value_checked(plan_data, "tddServiceDscpMap", index, issues, "ServiceMap", "P"),
        )
        set_row_font(worksheet, row_idx, times_font)


def process_rancm_expansion(plan_data, rancm_file):
    workbook = load_workbook_from_upload(rancm_file, data_only=False)
    issues = []

    sheet_actions = {
        "ManagedElement": lambda ws: fill_managed_element(ws, plan_data, issues),
        "Equipment": lambda ws: fill_equipment(ws, plan_data, issues),
        "RU": lambda ws: fill_ru(ws, plan_data, issues),
        "FiberDevice": lambda ws: fill_fiber_device(ws, plan_data, issues),
        "FiberCable": lambda ws: fill_fiber_cable(ws, plan_data, issues),
        "IrAntGroup": lambda ws: fill_ir_ant_group(ws, plan_data, issues),
        "IpLayerConfig": lambda ws: fill_ip_layer_config(ws, plan_data, issues),
        "Sctp": lambda ws: fill_sctp(ws) if is_cross_mode(plan_data) else None,
        "ServiceMap": lambda ws: fill_service_map(ws, plan_data, issues),
    }

    for sheet_name, action in sheet_actions.items():
        worksheet = get_sheet_by_name_fuzzy(workbook, sheet_name)
        if worksheet is None:
            issues.append(f"RANCM文件缺少Sheet[{sheet_name}]。")
            continue
        action(worksheet)

    sanitize_punctuation(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    summary = {
        "mode_type": get_mode_value(plan_data),
        "expansion_type": get_expansion_type_value(plan_data),
        "ru_count": len(get_ru_values(plan_data)),
    }
    return output, issues, summary


def fill_cfg_enbfunction(worksheet):
    existing_count = count_existing_rows(worksheet, 6, "A")
    target_count = max(existing_count, 1) + 1
    prepare_template_rows(worksheet, 6, target_count)

    base_row = 6
    target_row = 6 + max(existing_count, 1)
    base_snapshot = capture_row_snapshot(worksheet, base_row)
    write_row_snapshot(worksheet, target_row, base_snapshot)

    c_value = get_cell_value(worksheet, base_row, "C")
    d_value = get_cell_value(worksheet, base_row, "D")
    e_value = get_cell_value(worksheet, base_row, "E")
    set_cell_value(worksheet, target_row, "C", c_value)
    set_cell_value(worksheet, target_row, "D", d_value)
    set_cell_value(worksheet, target_row, "E", e_value)
    set_cell_value(worksheet, target_row, "H", e_value)
    set_cell_value(worksheet, target_row, "I", e_value)
    set_cell_value(worksheet, target_row, "J", d_value)
    set_cell_value(worksheet, target_row, "L", d_value)
    set_row_font(worksheet, target_row, times_font)


def fill_cfg_headers_by_target_headers(worksheet, row_idx, plan_data, issues, sheet_name, start_col, end_col, source_index):
    for col_idx in range(start_col, end_col + 1):
        header_value = worksheet.cell(row=1, column=col_idx).value
        if header_value is None or str(header_value).strip() == "":
            continue
        col_letter = get_column_letter(col_idx)
        value = get_plan_value_checked(plan_data, str(header_value).strip(), source_index, issues, sheet_name, col_letter)
        set_cell_value(worksheet, row_idx, col_letter, value)


def fill_cfg_tdd_cells(worksheet, plan_data, generated_ir_ant_numbers, issues):
    cell_count = get_cell_count(plan_data, issues, "Cell4GTDD")
    if cell_count <= 0:
        return

    existing_count = count_existing_rows(worksheet, 6, "A")
    target_total = existing_count + cell_count
    prepare_template_rows(worksheet, 6, target_total, max_col=64)

    cp_moid_seed = parse_int(get_cell_value(worksheet, 5 + existing_count, "AF"), 0)
    target_start_row = 6 + existing_count

    for index in range(cell_count):
        row_idx = target_start_row + index
        clear_row_columns(
            worksheet,
            row_idx,
            ["A", "H", "I", "L", "Q", "U", "V", "W", "Y", "AE", "AF", "AI", "AJ", "AK", "AL", "AM", "AN", "AO"],
        )
        set_cell_value(worksheet, row_idx, "A", "A")
        set_cell_value(worksheet, row_idx, "H", get_plan_value_checked(plan_data, "moId", index, issues, "Cell4GTDD", "H"))
        set_cell_value(worksheet, row_idx, "I", get_plan_value_checked(plan_data, "cellLocalId", index, issues, "Cell4GTDD", "I"))
        set_cell_value(worksheet, row_idx, "L", get_plan_value_checked(plan_data, "userLabel2", index, issues, "Cell4GTDD", "L"))
        set_cell_value(worksheet, row_idx, "Q", get_plan_value_checked(plan_data, "pci", index, issues, "Cell4GTDD", "Q"))
        set_cell_value(worksheet, row_idx, "U", get_plan_value_checked(plan_data, "freqBandInd", index, issues, "Cell4GTDD", "U"))
        set_cell_value(worksheet, row_idx, "V", get_plan_value_checked(plan_data, "earfcnUl", index, issues, "Cell4GTDD", "V"))
        set_cell_value(worksheet, row_idx, "W", get_plan_value_checked(plan_data, "bandWidthDl", index, issues, "Cell4GTDD", "W"))
        set_cell_value(worksheet, row_idx, "Y", get_plan_value_checked(plan_data, "bandWidthUl", index, issues, "Cell4GTDD", "Y"))
        set_cell_value(
            worksheet,
            row_idx,
            "AE",
            get_plan_value_checked(plan_data, "rootSequenceIndex", index, issues, "Cell4GTDD", "AE"),
        )
        set_cell_value(worksheet, row_idx, "AF", cp_moid_seed + index + 1)

        if index < len(generated_ir_ant_numbers):
            set_cell_value(worksheet, row_idx, "AI", f"{generated_ir_ant_numbers[index]}:1")
        else:
            issues.append(
                f"Sheet[Cell4GTDD] 列[AI] 参数[refIrAntGroup]：需要第{index + 1}行IrAntGroup数据，但生成的RANCM中不足。"
            )

        set_cell_value(
            worksheet,
            row_idx,
            "AJ",
            get_plan_value_checked(plan_data, "refBpDevice2", index, issues, "Cell4GTDD", "AJ"),
        )
        set_cell_value(worksheet, row_idx, "AK", get_plan_value_checked(plan_data, "cellMod", index, issues, "Cell4GTDD", "AK"))
        set_cell_value(
            worksheet,
            row_idx,
            "AL",
            get_plan_value_checked(plan_data, "cpSpeRefSigPwr", index, issues, "Cell4GTDD", "AL"),
        )
        set_cell_value(
            worksheet,
            row_idx,
            "AM",
            get_plan_value_checked(plan_data, "upActAntBitmapSeq", index, issues, "Cell4GTDD", "AM"),
        )
        set_cell_value(
            worksheet,
            row_idx,
            "AN",
            get_plan_value_checked(plan_data, "anttoPortMap", index, issues, "Cell4GTDD", "AN"),
        )
        set_cell_value(
            worksheet,
            row_idx,
            "AO",
            get_plan_value_checked(plan_data, "isDelNbrAndRelation", index, issues, "Cell4GTDD", "AO"),
        )
        set_row_font(worksheet, row_idx, times_font)


def fill_cfg_fdd_cells(worksheet, plan_data, issues):
    cell_count = get_cell_count(plan_data, issues, "Cell4GFDD")
    if cell_count <= 0:
        return

    existing_count = count_existing_rows(worksheet, 6, "A")
    target_total = existing_count + cell_count
    prepare_template_rows(worksheet, 6, target_total)

    target_start_row = 6 + existing_count
    for index in range(cell_count):
        row_idx = target_start_row + index
        clear_row_columns(
            worksheet,
            row_idx,
            ["A", "H", "I", "L", "Q", "S", "U", "V", "W", "X", "Y", "Z", "AA", "AF", "AG", "AH", "AI", "AK", "AL", "AM", "AP", "AQ", "AR", "AS", "AT", "AU", "AV"],
        )
        set_cell_value(worksheet, row_idx, "A", "A")
        set_cell_value(worksheet, row_idx, "H", get_plan_value_checked(plan_data, "moId", index, issues, "Cell4GFDD", "H"))
        set_cell_value(worksheet, row_idx, "I", get_plan_value_checked(plan_data, "cellLocalId", index, issues, "Cell4GFDD", "I"))
        set_cell_value(worksheet, row_idx, "L", get_plan_value_checked(plan_data, "userLabel2", index, issues, "Cell4GFDD", "L"))
        set_cell_value(worksheet, row_idx, "Q", get_plan_value_checked(plan_data, "pci", index, issues, "Cell4GFDD", "Q"))
        set_cell_value(worksheet, row_idx, "S", get_plan_value_checked(plan_data, "tac", index, issues, "Cell4GFDD", "S"))
        fill_cfg_headers_by_target_headers(worksheet, row_idx, plan_data, issues, "Cell4GFDD", column_index_from_string("U"), column_index_from_string("AA"), index)
        set_cell_value(worksheet, row_idx, "AF", get_plan_value_checked(plan_data, "LONGITUDE", index, issues, "Cell4GFDD", "AF"))
        set_cell_value(worksheet, row_idx, "AG", get_plan_value_checked(plan_data, "LATITUDE", index, issues, "Cell4GFDD", "AG"))
        set_cell_value(
            worksheet,
            row_idx,
            "AH",
            get_plan_value_checked(plan_data, "rootSequenceIndex", index, issues, "Cell4GFDD", "AH"),
        )
        set_cell_value(worksheet, row_idx, "AI", get_cell_value(worksheet, row_idx, "H"))
        set_cell_value(worksheet, row_idx, "AK", get_plan_value_checked(plan_data, "rfAppMode", index, issues, "Cell4GFDD", "AK"))
        fill_cfg_headers_by_target_headers(worksheet, row_idx, plan_data, issues, "Cell4GFDD", column_index_from_string("AL"), column_index_from_string("AM"), index)
        fill_cfg_headers_by_target_headers(worksheet, row_idx, plan_data, issues, "Cell4GFDD", column_index_from_string("AP"), column_index_from_string("AV"), index)
        set_row_font(worksheet, row_idx, times_font)


def process_cfg_radio_expansion(plan_data, cfg_file, generated_rancm_bytes):
    workbook = load_workbook_from_upload(cfg_file, data_only=False)
    issues = []
    mode_key = get_mode_key(plan_data)

    enb_sheet = get_sheet_by_name_fuzzy(workbook, "ENBFunction")
    if enb_sheet is None:
        issues.append("cfgRadioNet文件缺少Sheet[ENBFunction]。")
    else:
        fill_cfg_enbfunction(enb_sheet)

    if mode_key == "fddtotdd":
        target_sheet = get_sheet_by_name_fuzzy(workbook, "Cell4GTDD")
        if target_sheet is None:
            issues.append("cfgRadioNet文件缺少Sheet[Cell4GTDD]。")
        else:
            ir_ant_group_numbers = collect_generated_ir_ant_group_numbers(generated_rancm_bytes, issues)
            fill_cfg_tdd_cells(target_sheet, plan_data, ir_ant_group_numbers, issues)
    elif mode_key == "tddtofdd":
        target_sheet = get_sheet_by_name_fuzzy(workbook, "Cell4GFDD")
        if target_sheet is None:
            issues.append("cfgRadioNet文件缺少Sheet[Cell4GFDD]。")
        else:
            fill_cfg_fdd_cells(target_sheet, plan_data, issues)

    sanitize_punctuation(workbook)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, issues


def reset_expansion_state():
    for key in [
        "expansion_rancm_bytes",
        "expansion_cfg_bytes",
        "expansion_issues",
        "expansion_summary",
        "expansion_template_name",
        "expansion_rancm_name",
        "expansion_cfg_name",
    ]:
        st.session_state.pop(key, None)


def render_expansion_page():
    st.title("4G宏站扩容数据配置工具")
    st.markdown("---")
    st.info("先导入扩容参数模版表和RANCM；如果模板制式为跨制式，再额外导入cfgRadioNet文件。")

    template_file = st.file_uploader(
        "请导入扩容参数模版表",
        type=["xlsx"],
        key="expansion_template_file",
    )

    plan_data = None
    plan_error = None
    detected_mode = ""
    detected_expansion_type = ""
    requires_cfg_radio = False
    expected_cfg_label = ""

    if template_file is not None:
        try:
            plan_data = parse_plan_workbook(template_file)
            detected_mode = get_mode_value(plan_data)
            detected_expansion_type = get_expansion_type_value(plan_data)
            requires_cfg_radio = is_cross_mode(plan_data)
            if get_mode_key(plan_data) == "fddtotdd":
                expected_cfg_label = "RANCM-cfgRadioNet_TDD.xlsx"
            elif get_mode_key(plan_data) == "tddtofdd":
                expected_cfg_label = "RANCM-cfgRadioNet_FDD.xlsx"
        except Exception as exc:
            plan_error = str(exc)

    if template_file is not None:
        if plan_error:
            st.error(f"参数模版表解析失败：{plan_error}")
        else:
            st.success(f"已导入模版表：{template_file.name}")
            st.write(f"识别到扩容类型：`{detected_expansion_type or '未识别'}`")
            st.write(f"识别到制式：`{detected_mode or '未识别'}`")

    rancm_file = st.file_uploader(
        "请导入网管导出表 RANCM",
        type=["xlsx"],
        key="expansion_rancm_file",
    )
    if rancm_file is not None:
        st.success(f"已导入网管导出表：{rancm_file.name}")

    cfg_file = None
    if requires_cfg_radio:
        st.warning(f"当前是跨制式扩容，需要额外导入 `{expected_cfg_label}`。")
        cfg_file = st.file_uploader(
            f"请导入 {expected_cfg_label}",
            type=["xlsx"],
            key="expansion_cfg_file",
        )
        if cfg_file is not None:
            st.success(f"已导入cfgRadioNet表：{cfg_file.name}")

    current_template_name = template_file.name if template_file else None
    current_rancm_name = rancm_file.name if rancm_file else None
    current_cfg_name = cfg_file.name if cfg_file else None

    if (
        st.session_state.get("expansion_template_name") != current_template_name
        or st.session_state.get("expansion_rancm_name") != current_rancm_name
        or st.session_state.get("expansion_cfg_name") != current_cfg_name
    ):
        reset_expansion_state()
        st.session_state["expansion_template_name"] = current_template_name
        st.session_state["expansion_rancm_name"] = current_rancm_name
        st.session_state["expansion_cfg_name"] = current_cfg_name

    if st.button("开始生成扩容数据", key="expansion_generate_button"):
        if template_file is None:
            st.error("请先导入扩容参数模版表。")
        elif plan_error is not None or plan_data is None:
            st.error("参数模版表无法解析，请先修正后再生成。")
        elif rancm_file is None:
            st.error("请先导入网管导出表 RANCM。")
        elif requires_cfg_radio and cfg_file is None:
            st.error(f"当前制式为跨制式扩容，请先导入 {expected_cfg_label}。")
        else:
            try:
                with st.spinner("正在生成扩容结果，请稍候..."):
                    rancm_output, rancm_issues, summary = process_rancm_expansion(plan_data, rancm_file)
                    cfg_output = None
                    cfg_issues = []
                    if requires_cfg_radio and cfg_file is not None:
                        cfg_output, cfg_issues = process_cfg_radio_expansion(plan_data, cfg_file, rancm_output.getvalue())

                st.session_state["expansion_rancm_bytes"] = rancm_output.getvalue()
                st.session_state["expansion_cfg_bytes"] = cfg_output.getvalue() if cfg_output is not None else None
                st.session_state["expansion_issues"] = rancm_issues + cfg_issues
                st.session_state["expansion_summary"] = summary
            except Exception as exc:
                st.error(f"扩容数据生成失败：{exc}")
                st.exception(exc)

    rancm_bytes = st.session_state.get("expansion_rancm_bytes")
    cfg_bytes = st.session_state.get("expansion_cfg_bytes")
    cfg_download_name = f"Result_{expected_cfg_label}" if expected_cfg_label else "Result_RANCM-cfgRadioNet.xlsx"
    if rancm_bytes:
        summary = st.session_state.get("expansion_summary", {})
        st.success("扩容数据生成完成。")
        st.write(
            f"扩容类型：`{summary.get('expansion_type', '')}` | "
            f"制式：`{summary.get('mode_type', '')}` | "
            f"RU数量：`{summary.get('ru_count', 0)}`"
        )

        if cfg_bytes:
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="下载扩容后的 RANCM",
                    data=rancm_bytes,
                    file_name="Result_RANCM_Expansion.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="expansion_download_rancm_button",
                )
            with col2:
                st.download_button(
                    label="下载扩容后的 cfgRadioNet",
                    data=cfg_bytes,
                    file_name=cfg_download_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="expansion_download_cfg_button",
                )
        else:
            st.download_button(
                label="下载扩容后的 RANCM",
                data=rancm_bytes,
                file_name="Result_RANCM_Expansion.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="expansion_download_rancm_only_button",
            )

        issues = st.session_state.get("expansion_issues", [])
        if issues:
            st.warning(f"本次生成有 {len(issues)} 条提示，请核对。")
            with st.expander("查看处理提示", expanded=False):
                for issue in issues:
                    st.write(f"- {issue}")


if __name__ == "__main__":
    st.set_page_config(page_title="SDR 扩容数据配置工具", layout="centered")
    render_expansion_page()
