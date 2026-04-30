import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
from io import BytesIO
import os

# === 配置区域 ===
# 必须确保同目录下有该模版文件，用于提取标准参数
TEMPLATE_FILENAME = "chongxuan.xlsx"

# === 定义全局字体样式 ===
times_font = Font(name='Times New Roman')


def get_sheet_by_name(wb, sheet_name):
    """
    安全获取 Sheet 的辅助函数。
    """
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    else:
        # 尝试去除空格匹配
        for name in wb.sheetnames:
            if name.strip() == sheet_name.strip():
                return wb[name]
        raise ValueError(f"错误：上传的文件中找不到名称为 '{sheet_name}' 的工作表，请检查文件格式。")


def load_template_params(template_path):
    """
    从内置模版中读取 Sheet6 (GeranMeasFDDLTE) 和 Sheet8 (GsmRslParaFDD) 的标准参数。
    """
    params = {
        "sheet6": {},
        "sheet8": {}
    }

    if not os.path.exists(template_path):
        return None, f"未找到内置模版文件: {template_path}"

    try:
        wb = openpyxl.load_workbook(template_path, data_only=True)

        # --- 读取 Sheet 6 模版参数 ---
        if "GeranMeasFDDLTE" in wb.sheetnames:
            ws6 = wb["GeranMeasFDDLTE"]
            # 读取 H(8) 到 AY(51) 列的第6行数据
            for col in range(8, 52):
                val = ws6.cell(row=6, column=col).value
                params["sheet6"][col] = val

        # --- 读取 Sheet 8 模版参数 ---
        if "GsmRslParaFDD" in wb.sheetnames:
            ws8 = wb["GsmRslParaFDD"]
            # 读取 H(8) 到 AX(50) 列的第6行数据
            for col in range(8, 51):
                val = ws8.cell(row=6, column=col).value
                params["sheet8"][col] = val

        wb.close()
        return params, None
    except Exception as e:
        return None, f"读取模版出错: {str(e)}"


def process_excel(uploaded_file, template_params):
    """
    核心处理逻辑
    """
    wb = openpyxl.load_workbook(uploaded_file)

    # =========================================================
    # 步骤 1: 操作 Sheet 3 [CUEUtranCellFDDLTE]
    # =========================================================
    ws3 = get_sheet_by_name(wb, "CUEUtranCellFDDLTE")
    col_O = column_index_from_string('O')
    col_AW = column_index_from_string('AW')
    col_GK = column_index_from_string('GK')

    for row in ws3.iter_rows(min_row=6, max_row=ws3.max_row):
        if not row[0].value and not row[1].value:
            continue

        updates = [
            (col_O, "容量自适应[4]"),
            (col_AW, "解关断[0]"),
            (col_GK, "打开[1]"),
            (1, "M")  # A列
        ]

        for col_idx, val in updates:
            cell = ws3.cell(row=row[0].row, column=col_idx)
            cell.value = val
            cell.font = times_font

    # =========================================================
    # 步骤 2: 操作 Sheet 4 [EUtranCellMeasFDDLTE]
    # 【关键更正】：此表既要修改自身，又是 Sheet 6 的数据源
    # =========================================================
    ws4 = get_sheet_by_name(wb, "EUtranCellMeasFDDLTE")
    col_AH = column_index_from_string('AH')

    # 用于 Sheet 6 的数据源
    data_for_sheet6 = []

    for row in ws4.iter_rows(min_row=6, max_row=ws4.max_row):
        # 1. 提取数据供 Sheet 6 使用 (提取 B-F 列)
        if row[1].value is not None:
            # 注意：row[1] 是 B 列
            row_data = [row[1].value, row[2].value, row[3].value, row[4].value, row[5].value]
            data_for_sheet6.append(row_data)

        # 2. 修改 Sheet 4 自身
        if not row[0].value and not row[1].value:
            continue

        # A列 -> M
        cell_a = ws4.cell(row=row[0].row, column=1)
        cell_a.value = "M"
        cell_a.font = times_font

        # AH列 -> 1
        cell_ah = ws4.cell(row=row[0].row, column=col_AH)
        cell_ah.value = 1
        cell_ah.font = times_font

    # =========================================================
    # 步骤 3: 操作 Sheet 5 [RatPriCnParaFDDLTE]
    # =========================================================
    ws5 = get_sheet_by_name(wb, "RatPriCnParaFDDLTE")
    col_H = column_index_from_string('H')
    col_I = column_index_from_string('I')

    for row in ws5.iter_rows(min_row=6, max_row=ws5.max_row):
        if row[1].value is None:
            continue

        c_h = ws5.cell(row=row[0].row, column=col_H)
        c_h.value = "100"
        c_h.font = times_font

        c_i = ws5.cell(row=row[0].row, column=col_I)
        c_i.value = "0"
        c_i.font = times_font

    # =========================================================
    # 步骤 4: 操作 Sheet 6 [GeranMeasFDDLTE]
    # 【数据源】：Sheet 4 (EUtranCellMeasFDDLTE)
    # =========================================================
    ws6 = get_sheet_by_name(wb, "GeranMeasFDDLTE")
    start_row = 6

    for i, data in enumerate(data_for_sheet6):
        current_row = start_row + i

        # A列
        ws6.cell(row=current_row, column=1).value = "A"

        # B-E列 (从 Sheet 4 复制)
        ws6.cell(row=current_row, column=2).value = data[0]
        ws6.cell(row=current_row, column=3).value = data[1]
        ws6.cell(row=current_row, column=4).value = data[2]
        ws6.cell(row=current_row, column=5).value = data[3]

        # F列 (内容 + ,GeranMeasFDDLTE=1)
        original_dn = data[4] if data[4] else ""
        ws6.cell(row=current_row, column=6).value = f"{original_dn},GeranMeasFDDLTE=1"

        # H-AY 列 (从模版 Sheet 6 复制)
        for col in range(8, 52):
            if col in template_params['sheet6']:
                ws6.cell(row=current_row, column=col).value = template_params['sheet6'][col]

        # 统一字体
        for col in range(1, 52):
            cell = ws6.cell(row=current_row, column=col)
            cell.font = times_font

    # =========================================================
    # 步骤 5: 操作 Sheet 7 [GsmReselectionFDDLTE]
    # 【关键更正】：此表既要修改自身，又是 Sheet 8 的数据源
    # =========================================================
    ws7 = get_sheet_by_name(wb, "GsmReselectionFDDLTE")
    col_J = column_index_from_string('J')
    col_K = column_index_from_string('K')
    col_L = column_index_from_string('L')

    # 用于 Sheet 8 的数据源
    data_for_sheet8 = []

    for row in ws7.iter_rows(min_row=6, max_row=ws7.max_row):
        # 1. 提取数据供 Sheet 8 使用 (B-F列)
        if row[1].value is not None:
            row_vals = [row[1].value, row[2].value, row[3].value, row[4].value, row[5].value]
            data_for_sheet8.append(row_vals)

        # 2. 修改 Sheet 7 自身
        if not row[0].value: continue

        updates = [
            (col_J, "1[3]"),
            (col_K, "1[3]"),
            (col_L, 1)
        ]

        for col_idx, val in updates:
            cell = ws7.cell(row=row[0].row, column=col_idx)
            cell.value = val
            cell.font = times_font

    # =========================================================
    # 步骤 6: 操作 Sheet 8 [GsmRslParaFDD]
    # 【数据源】：Sheet 7 (GsmReselectionFDDLTE)
    # =========================================================
    ws8 = get_sheet_by_name(wb, "GsmRslParaFDD")

    for i, data in enumerate(data_for_sheet8):
        current_row = start_row + i

        # A列
        ws8.cell(row=current_row, column=1).value = "A"

        # B-E列 (从 Sheet 7 复制)
        ws8.cell(row=current_row, column=2).value = data[0]
        ws8.cell(row=current_row, column=3).value = data[1]
        ws8.cell(row=current_row, column=4).value = data[2]
        ws8.cell(row=current_row, column=5).value = data[3]

        # F列 (内容 + ,GsmRslParaFDD=1)
        original_dn = data[4] if data[4] else ""
        ws8.cell(row=current_row, column=6).value = f"{original_dn},GsmRslParaFDD=1"

        # H-AX 列 (从模版 Sheet 8 复制)
        for col in range(8, 51):
            if col in template_params['sheet8']:
                ws8.cell(row=current_row, column=col).value = template_params['sheet8'][col]

        # 统一字体
        for col in range(1, 51):
            cell = ws8.cell(row=current_row, column=col)
            cell.font = times_font

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# === Streamlit 界面逻辑 ===

st.set_page_config(page_title="ITBBU 自动化配置工具", layout="centered")

st.title("🛠️ ITBBU Excel 自动化配置工具")
st.info("请确保 `template.xlsx` (内置参数模版) 已放置在工具同级目录下。")
st.markdown("""
**目前是ITBBU的FDD重新数据制作流程：**
*还剩ITBBU的TDD，SDR的FDD、TDD。代码逻辑类似
""")

template_params, error_msg = load_template_params(TEMPLATE_FILENAME)

if error_msg:
    st.error(f"⚠️ 初始化失败: {error_msg}")
else:
    st.success(f"✅ 系统就绪 | 模版已加载 | 字体: Times New Roman")

    uploaded_file = st.file_uploader("请导入 ITBBU 重选模版 (xlsx)", type=["xlsx"])

    if uploaded_file is not None:
        if st.button("开始自动化处理"):
            try:
                processed_data = process_excel(uploaded_file, template_params)
                st.success("处理完成！")
                st.download_button(
                    label="📥 下载处理后的文件",
                    data=processed_data,
                    file_name="ITBBU重选模版_已处理.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except ValueError as ve:
                # 捕获已知错误（如找不到Sheet）
                st.error(f"数据格式错误: {str(ve)}")
            except Exception as e:
                # 捕获未知错误，并打印出来供排查
                st.error(f"发生了未知错误: {str(e)}")