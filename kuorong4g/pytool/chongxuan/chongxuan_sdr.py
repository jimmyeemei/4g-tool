import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font
from io import BytesIO
import re

# ==============================================================================
# 1. 初始化页面配置
# ==============================================================================
st.set_page_config(
    page_title="SDR 重选模版制作工具",
    layout="centered"
)

# 定义统一字体
times_font = Font(name='Times New Roman')


# ==============================================================================
# 2. 核心辅助函数
# ==============================================================================

def normalize_key(key):
    """
    标准化表头：去除空格、下划线，转小写，用于模糊匹配
    """
    if not key:
        return ""
    # 只保留字母和数字
    return re.sub(r'[^a-zA-Z0-9]', '', str(key)).lower()


def find_column_index_by_header(ws, header_name, search_rows=[1]):
    """
    根据表头名称查找列号（1-based index）
    :param search_rows: 列表，指定在哪些行查找表头（通常是第1行，有时是前5行）
    """
    target = normalize_key(header_name)
    for r in search_rows:
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r, column=col).value
            if cell_val and normalize_key(cell_val) == target:
                return col
    return None


def get_valid_row_count(ws, start_row=6):
    """
    根据 'NE_Name' 列计算有效数据行数
    """
    # 优先找 NE_Name，找不到找 ManagedElement
    ne_name_col = find_column_index_by_header(ws, "NE_Name", search_rows=[1, 2, 3, 4, 5])
    if not ne_name_col:
        ne_name_col = find_column_index_by_header(ws, "ManagedElement", search_rows=[1, 2, 3, 4, 5])

    # 如果都找不到，且最大行数大于起始行，则默认全部处理（兜底逻辑）
    if not ne_name_col:
        return max(0, ws.max_row - start_row + 1)

    count = 0
    # 从起始行开始向下遍历，直到遇到空值
    for r in range(start_row, ws.max_row + 100):  # 多读一些缓冲行
        val = ws.cell(row=r, column=ne_name_col).value
        if val is not None and str(val).strip() != "":
            count += 1
        else:
            break
    return count


def set_cell_value(ws, row, col, value):
    """设置单元格值并应用字体"""
    if col is None:
        return  # 如果列没找到，跳过
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.font = times_font


# ==============================================================================
# 3. 业务处理逻辑
# ==============================================================================

def process_sdr_chongxuan(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    logs = []

    # -------------------------------------------------------------------------
    # Sheet 1: EUtranCellFDD
    # -------------------------------------------------------------------------
    sheet_name = "EUtranCellFDD"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        valid_rows = get_valid_row_count(ws, 6)
        logs.append(f"✅ [{sheet_name}] 识别到有效数据行数: {valid_rows} 行")

        if valid_rows > 0:
            # 查找列索引 (在第1行查找表头)
            col_capa = find_column_index_by_header(ws, "cellCapaLeveInd")
            col_admin = find_column_index_by_header(ws, "adminState")
            col_mod = find_column_index_by_header(ws, "MODIND")

            for i in range(valid_rows):
                r = 6 + i
                if col_capa: set_cell_value(ws, r, col_capa, "容量自适应[4]")
                if col_admin: set_cell_value(ws, r, col_admin, "解关断[0]")
                # 修改点：MODIND 改为 M
                if col_mod: set_cell_value(ws, r, col_mod, "M")
    else:
        logs.append(f"⚠️ 未找到 Sheet: {sheet_name}")

    # -------------------------------------------------------------------------
    # Sheet 2: EUtranCellMeasurement
    # -------------------------------------------------------------------------
    sheet_name = "EUtranCellMeasurement"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        valid_rows = get_valid_row_count(ws, 6)
        logs.append(f"✅ [{sheet_name}] 识别到有效数据行数: {valid_rows} 行")

        if valid_rows > 0:
            # 1. 查找列索引
            col_mod = find_column_index_by_header(ws, "MODIND")  # 新增需求
            col_csfb2 = find_column_index_by_header(ws, "ratPriCnCSFB2")
            # 注意：表头名字较长，确保 excel 中第一行确实是这个名字
            col_para_csfb1 = find_column_index_by_header(ws, "ratPriCnPara_ratPriCnCSFB1")
            col_geran = find_column_index_by_header(ws, "geranCarriFreqNum")
            col_para_csfb2 = find_column_index_by_header(ws, "ratPriCnPara_ratPriCnCSFB2")

            for i in range(valid_rows):
                r = 6 + i
                if col_mod: set_cell_value(ws, r, col_mod, "M")
                if col_csfb2: set_cell_value(ws, r, col_csfb2, 0)
                if col_para_csfb1: set_cell_value(ws, r, col_para_csfb1, 100)
                if col_geran: set_cell_value(ws, r, col_geran, 1)
                if col_para_csfb2: set_cell_value(ws, r, col_para_csfb2, 0)

            # 2. 批量填充 CS 到 EI 列 (共43个参数)
            params_cs_to_ei = [
                "4[0]", "关闭[0]", "关闭[0]", 16, 1, "打开[1]", "关闭[0]", "1[0]",
                "关闭[0]", "关闭[0]", "关闭[0]", "关闭[0]", "关闭[0]", "关闭[0]", "关闭[0]", "关闭[0]",
                "关闭[0]", "关闭[0]", 0, 0, 0, 0, 12, 12,
                10, 10, "关闭[0]", "打开[1]", "时域算法[1]", 1800, "300公里/小时[2]", 10,
                10, "关闭[0]", "关闭[0]", -70, 100, 100, 2, 58,
                "关闭[0]", "关闭[0]", "关闭[0]"
            ]

            start_col_idx = column_index_from_string("CS")

            for i in range(valid_rows):
                r = 6 + i
                for p_idx, param in enumerate(params_cs_to_ei):
                    # 计算当前列号
                    current_col = start_col_idx + p_idx
                    set_cell_value(ws, r, current_col, param)
    else:
        logs.append(f"⚠️ 未找到 Sheet: {sheet_name}")

    # -------------------------------------------------------------------------
    # Sheet 3: GsmReselection
    # -------------------------------------------------------------------------
    sheet_name = "GsmReselection"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        valid_rows = get_valid_row_count(ws, 6)
        logs.append(f"✅ [{sheet_name}] 识别到有效数据行数: {valid_rows} 行")

        if valid_rows > 0:
            # 1. 修改特定表头参数
            col_med = find_column_index_by_header(ws, "sfMediumGERAN")
            col_high = find_column_index_by_header(ws, "sfHighGERAN")
            col_num = find_column_index_by_header(ws, "geranFreqNum")

            for i in range(valid_rows):
                r = 6 + i
                if col_med: set_cell_value(ws, r, col_med, "1[3]")
                if col_high: set_cell_value(ws, r, col_high, "1[3]")
                if col_num: set_cell_value(ws, r, col_num, 1)

                # A列填写 M
                set_cell_value(ws, r, 1, "M")

            # 2. 批量填充 M 到 BB 列 (共42个参数)
            params_m_to_bb = [
                0, 0, 1, 0, -99, 23, 10, 14,
                1, 255, 46, 47, 48, 49, 50, 51,
                52, 53, 54, 55, 56, 57, 59, 60,
                61, 62, 63, 64, 512, 513, 514, 515,
                516, 517, 518, 522, 524, 526, 528, 531,
                533, 534
            ]

            start_col_idx = column_index_from_string("M")

            for i in range(valid_rows):
                r = 6 + i
                for p_idx, param in enumerate(params_m_to_bb):
                    current_col = start_col_idx + p_idx
                    set_cell_value(ws, r, current_col, param)
    else:
        logs.append(f"⚠️ 未找到 Sheet: {sheet_name}")

    # 保存文件
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, logs


# ==============================================================================
# 4. 主界面逻辑
# ==============================================================================

st.title("🛠️ SDR 重选数据制作工具")
st.markdown("---")

st.info("请上传名为 **'SDR重选模版.xlsx'** (或包含对应Sheet) 的文件。")

uploaded_file = st.file_uploader("📂 上传文件", type=["xlsx"])

if uploaded_file:
    if st.button("🚀 开始处理"):
        try:
            with st.spinner("正在处理数据，请稍候..."):
                processed_data, logs = process_sdr_chongxuan(uploaded_file)

            st.success("处理完成！")

            # 显示日志
            with st.expander("查看处理详情日志"):
                for log in logs:
                    st.write(log)

            # 下载按钮
            st.download_button(
                label="📥 下载处理后的结果文件",
                data=processed_data,
                file_name="SDR_Reselection_Result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"❌ 处理过程中发生错误: {str(e)}")