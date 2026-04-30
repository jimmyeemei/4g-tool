import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import Font
from io import BytesIO
import os
import re
from copy import copy

# ==============================================================================
# 1. 初始化
# ==============================================================================
st.set_page_config(
    page_title="SDR 开站数据配置工具",
    layout="centered"
)

# 模版文件定义
TEMPLATE_RANCM = "RANCM.xlsx"
TEMPLATE_FDD = "cfgradioFDD.xlsx"
TEMPLATE_TDD = "cfgradioTDD.xlsx"

SDR_SOURCE_SHEET_NAME = "RANCM-sdrPlan"
times_font = Font(name='Times New Roman')


# ==============================================================================
# 2. 辅助函数
# ==============================================================================

def normalize_key(key):
    if not key:
        return ""
    return re.sub(r'[^a-zA-Z0-9]', '', str(key)).lower()


def get_sheet_by_name_fuzzy(wb, target_name):
    if target_name in wb.sheetnames:
        return wb[target_name]
    target_lower = target_name.lower().strip()
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() == target_lower:
            return wb[sheet_name]
    return None


def set_row_font(ws, row_idx, font):
    for cell in ws[row_idx]:
        cell.font = font


def sanitize_punctuation(wb):
    """
    【全局标点净化引擎】
    在文件保存前，扫描整个工作簿，将所有中文字符标点强行转换为英文字符标点，防止导入基站报错。
    """
    punct_map = str.maketrans(
        "，。：；！？（）【】“”‘’",
        ",.:;!?()[]\"\"''"
    )
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = cell.value.translate(punct_map)


def prepare_template_rows(ws, start_row, target_count):
    """
    【光速版按列独立校验引擎】
    完全遵循：定量数据不足n行则同列内向下复制填充，等于n行不操作，大于n行删去多余行。
    """
    max_col = min(ws.max_column, 150)

    for c in range(1, max_col + 1):
        last_valid_r = start_row - 1

        for r in range(start_row, ws.max_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                last_valid_r = r
            else:
                break

        if last_valid_r < start_row:
            last_valid_r = start_row

        current_count = last_valid_r - start_row + 1

        if current_count < target_count:
            source_cell = ws.cell(row=last_valid_r, column=c)
            src_val = source_cell.value
            src_has_style = source_cell.has_style
            src_style = source_cell._style if src_has_style else None

            for r in range(last_valid_r + 1, start_row + target_count):
                target_cell = ws.cell(row=r, column=c)
                target_cell.value = src_val
                if src_has_style:
                    target_cell._style = src_style

    valid_max_row = start_row + target_count - 1
    if ws.max_row > valid_max_row:
        ws.delete_rows(valid_max_row + 1, ws.max_row - valid_max_row)


def get_sdr_source_data(uploaded_file):
    try:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws = get_sheet_by_name_fuzzy(wb, SDR_SOURCE_SHEET_NAME)
        if not ws:
            return None, None, f"未找到Sheet: {SDR_SOURCE_SHEET_NAME}"

        headers = {}
        header_debug = []
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                raw_header = str(cell.value).strip()
                headers[col_idx] = raw_header
                header_debug.append(raw_header)

        raw_data_map = {name: [] for name in headers.values()}
        normalized_key_map = {normalize_key(k): k for k in headers.values()}

        max_row = ws.max_row
        for col_idx, name in headers.items():
            for r in range(4, max_row + 1):
                val = ws.cell(row=r, column=col_idx).value
                raw_data_map[name].append(val if val is not None else "")

        ws_rru = get_sheet_by_name_fuzzy(wb, "RRU")
        rru_grid = {}
        rru_max_row = 0
        rru_max_col = 0
        rru_search_index = {}

        if ws_rru:
            rru_max_row = min(ws_rru.max_row, 20)
            rru_max_col = ws_rru.max_column
            for r in range(1, rru_max_row + 1):
                rru_grid[r] = {}
                for c in range(1, rru_max_col + 1):
                    val = ws_rru.cell(row=r, column=c).value
                    rru_grid[r][c] = val
                    if r >= 3 and val is not None and str(val).strip() != "":
                        clean_val = str(val).strip()
                        if clean_val not in rru_search_index:
                            rru_search_index[clean_val] = (r, c)

        return (raw_data_map, normalized_key_map, rru_grid, rru_max_row, rru_max_col,
                rru_search_index), header_debug, None
    except Exception as e:
        return None, None, str(e)


def get_val_sdr_strict(source_data_pack, template_key, target_idx, error_context=None):
    """
    【智能降级数据读取引擎】
    """
    raw_data_map, normalized_key_map = source_data_pack[:2]
    clean_key = normalize_key(template_key)

    if clean_key in normalized_key_map:
        real_key = normalized_key_map[clean_key]
        values = list(raw_data_map[real_key])
        while values and values[-1] == "":
            values.pop()
    else:
        return "", None

    if not values:
        return "", None

    count = len(values)

    if count == 1:
        return values[0], None

    if target_idx < count:
        return values[target_idx], None
    else:
        allow_empty_keys = ["vid", "ipaddr", "networkmask", "gatewayip", "moid"]
        auto_fill_keys = ["rfappmode", "refbpdevice", "refbpdevice2", "cpsperefsigpwr", "maxcptranspwr"]

        if clean_key in auto_fill_keys:
            return values[-1], None

        if clean_key in allow_empty_keys:
            return "", None

        if error_context:
            sheet = error_context.get('sheet', 'Unknown')
            col = error_context.get('col', 'Unknown')
            return "ERROR", f"Sheet[{sheet}] 列[{col}] 参数[{template_key}]: 需要第{target_idx + 1}行数据，但源数据只有{count}行。"

        return "", None


# ==============================================================================
# 3. 核心处理逻辑 - RANCM
# ==============================================================================

def process_sdr_rancm(template_path, source_data_pack, mode="FDD"):
    if not os.path.exists(template_path):
        return None, [f"本地模版缺失: {template_path}"]

    wb = openpyxl.load_workbook(template_path)
    errors = []

    try:
        val, _ = get_val_sdr_strict(source_data_pack, "cellnum", 0)
        cell_num = int(val) if val else 1
    except:
        cell_num = 1

    try:
        val_numofB, _ = get_val_sdr_strict(source_data_pack, "numofB", 0)
        if not val_numofB: val_numofB = "12"
    except:
        val_numofB = "12"

    raw_map, norm_map = source_data_pack[:2]
    ru_key = norm_map.get(normalize_key("RUDevice"))
    ru_rows = 1
    if ru_key:
        ru_values = [v for v in raw_map[ru_key] if v != ""]
        if ru_values:
            ru_rows = len(ru_values)

    def fill_base_cols(ws, row, sheet_name):
        ws[f"A{row}"] = "A"
        ws[f"B{row}"] = "SDR"
        v_sub, err1 = get_val_sdr_strict(source_data_pack, "SubNetwork", 0, {'sheet': sheet_name, 'col': 'C'})
        v_me, err2 = get_val_sdr_strict(source_data_pack, "ManagedElement", 0, {'sheet': sheet_name, 'col': 'D'})
        v_ne, err3 = get_val_sdr_strict(source_data_pack, "NE_Name", 0, {'sheet': sheet_name, 'col': 'E'})
        if err1: errors.append(err1)
        if err2: errors.append(err2)
        if err3: errors.append(err3)
        ws[f"C{row}"] = v_sub
        ws[f"D{row}"] = v_me
        ws[f"E{row}"] = v_ne

    # --- 1. ManagedElement ---
    s_name = "ManagedElement"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        r = 6
        count = 1
        prepare_template_rows(ws, 6, count)
        fill_base_cols(ws, r, s_name)
        mapping = {
            "F": "mimType",
            "G": "mimVersion", "H": "RADIOMODE", "I": "SWVERSION",
            "M": "MEADDR", "Q": "NE_Name", "R": "NE_Name",
            "Y": "LONGITUDE", "Z": "LATITUDE", "AR": "RUNRADIOMODE"
        }
        for col, key in mapping.items():
            val, err = get_val_sdr_strict(source_data_pack, key, 0, {'sheet': s_name, 'col': col})
            if err:
                errors.append(err)
            else:
                ws[f"{col}{r}"] = val
        set_row_font(ws, r, times_font)

    # --- 2. Equipment ---
    s_name = "Equipment"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    equip_cache_B_E = []
    if ws:
        r = 6
        count = 1
        prepare_template_rows(ws, 6, count)
        fill_base_cols(ws, r, s_name)
        for col_idx in range(6, 36):
            col_letter = get_column_letter(col_idx)
            h_val = ws.cell(row=1, column=col_idx).value
            if h_val:
                key = str(h_val).strip()
                val, err = get_val_sdr_strict(source_data_pack, key, 0, {'sheet': s_name, 'col': col_letter})
                if err:
                    errors.append(err)
                elif val != "":
                    ws.cell(row=r, column=col_idx).value = val

        set_row_font(ws, r, times_font)
        for c in range(2, 6):
            equip_cache_B_E.append(ws.cell(row=r, column=c).value)

    # --- 3. RU ---
    s_name = "RU"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = ru_rows
        prepare_template_rows(ws, 6, count)
        for i in range(count):
            curr = 6 + i
            fill_base_cols(ws, curr, s_name)
            params = [("F", "RUDevice"), ("G", "userLabel1"), ("H", "RUType"),
                      ("K", "RADIOMODE"), ("L", "functionMode"), ("M", "RUDevice")]
            for col, key in params:
                val, err = get_val_sdr_strict(source_data_pack, key, i, {'sheet': s_name, 'col': col})
                if err:
                    errors.append(err)
                else:
                    ws[f"{col}{curr}"] = val
            set_row_font(ws, curr, times_font)

    # --- 4. FiberCable ---
    s_name = "FiberCable"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = cell_num
        prepare_template_rows(ws, 6, count)

        rru_grid = source_data_pack[2]
        rru_search_index = source_data_pack[5]

        for i in range(count):
            curr = 6 + i
            if len(equip_cache_B_E) >= 4:
                for idx, val in enumerate(equip_cache_B_E):
                    ws.cell(row=curr, column=idx + 2).value = val

            f_val, err = get_val_sdr_strict(source_data_pack, "RUDevice", i, {'sheet': s_name, 'col': 'F'})
            if err:
                errors.append(err)
            else:
                ws[f"F{curr}"] = f_val

            found_r = None
            found_c = None

            if f_val is not None and str(f_val).strip() != "":
                target_val = str(f_val).strip()
                if target_val in rru_search_index:
                    found_r, found_c = rru_search_index[target_val]

            if found_r == 3:
                x = rru_grid.get(1, {}).get(found_c) if rru_grid.get(1, {}).get(found_c) is not None else ""
                y = rru_grid.get(found_r - 1, {}).get(found_c) if rru_grid.get(found_r - 1, {}).get(
                    found_c) is not None else ""
                ws[f"G{curr}"] = f"(1,1,{x}):{y}"
            elif found_r and found_r > 3:
                z = rru_grid.get(found_r - 1, {}).get(found_c) if rru_grid.get(found_r - 1, {}).get(
                    found_c) is not None else ""
                ws[f"G{curr}"] = f"({z},1,1):2"
            else:
                ws[f"G{curr}"] = f"(1,1,{val_numofB}):{i}"

            if found_r:
                next_val = rru_grid.get(found_r + 1, {}).get(found_c)
                if next_val is not None and str(next_val).strip() != "":
                    ws[f"H{curr}"] = f"({f_val},1,1):1"
                else:
                    ws[f"H{curr}"] = f"({f_val},1,1):1"
            else:
                ws[f"H{curr}"] = f"({f_val},1,1):1"

            set_row_font(ws, curr, times_font)

    # --- 5. PhyLayerPort ---
    s_name = "PhyLayerPort"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        r = 6
        count = 1
        prepare_template_rows(ws, 6, count)

        col_radio = None
        col_proto = None
        if mode == "TDD":
            for c in range(1, ws.max_column + 1):
                h_val = str(ws.cell(row=1, column=c).value or "").strip().lower()
                if h_val == "radiomode":
                    col_radio = c
                elif h_val == "protocoltype":
                    col_proto = c

        fill_base_cols(ws, r, s_name)
        val, err = get_val_sdr_strict(source_data_pack, "refGeDevice", 0, {'sheet': s_name, 'col': 'K'})
        if err:
            errors.append(err)
        else:
            ws[f"K{r}"] = val

        if mode == "TDD":
            if col_radio: ws.cell(row=r, column=col_radio).value = 32
            if col_proto: ws.cell(row=r, column=col_proto).value = "PHY LTE IR[1]"

        set_row_font(ws, r, times_font)

    # --- 6. IpLayerConfig ---
    s_name = "IpLayerConfig"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = 2
        prepare_template_rows(ws, 6, count)
        for i in range(count):
            curr = 6 + i
            fill_base_cols(ws, curr, s_name)
            for col_idx in range(10, 16):
                col_letter = get_column_letter(col_idx)
                h_val = ws.cell(row=1, column=col_idx).value
                if h_val:
                    key = str(h_val).strip()
                    val, err = get_val_sdr_strict(source_data_pack, key, i, {'sheet': s_name, 'col': col_letter})
                    if err:
                        errors.append(err)
                    elif val != "":
                        ws.cell(row=curr, column=col_idx).value = val
            set_row_font(ws, curr, times_font)

    # --- 7. Sctp ---
    s_name = "Sctp"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = 10
        prepare_template_rows(ws, 6, count)

        col_radio = None
        if mode == "TDD":
            for c in range(1, ws.max_column + 1):
                h_val = str(ws.cell(row=1, column=c).value or "").strip().lower()
                if h_val == "radiomode":
                    col_radio = c
                    break

        for i in range(count):
            curr = 6 + i
            fill_base_cols(ws, curr, s_name)
            ws[f"G{curr}"] = i + 1
            ws[f"H{curr}"] = i + 1

            if mode == "TDD" and col_radio:
                ws.cell(row=curr, column=col_radio).value = 32

            set_row_font(ws, curr, times_font)

    # --- 8. ServiceMap ---
    s_name = "ServiceMap"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = 2
        prepare_template_rows(ws, 6, count)
        for i in range(count):
            curr = 6 + i
            fill_base_cols(ws, curr, s_name)
            ws[f"F{curr}"] = i + 1
            for col_idx in range(15, 18):
                col_letter = get_column_letter(col_idx)
                h_val = ws.cell(row=1, column=col_idx).value
                if h_val:
                    key = str(h_val).strip()
                    val, err = get_val_sdr_strict(source_data_pack, key, i, {'sheet': s_name, 'col': col_letter})
                    if err:
                        errors.append(err)
                    elif val != "":
                        ws.cell(row=curr, column=col_idx).value = val
            set_row_font(ws, curr, times_font)

    # --- 9. FiberDevice ---
    s_name = "FiberDevice"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        fd_rows = ru_rows
        prepare_template_rows(ws, 6, fd_rows)

        col_proto = None
        if mode == "TDD":
            for c in range(1, ws.max_column + 1):
                h_val = str(ws.cell(row=1, column=c).value or "").strip().lower()
                if h_val == "protocoltype":
                    col_proto = c
                    break

        for i in range(fd_rows):
            curr = 6 + i
            ws[f"A{curr}"] = "M"
            ws[f"B{curr}"] = "SDR"
            for col_letter in ['C', 'D', 'E']:
                col_idx = column_index_from_string(col_letter)
                h_val = ws.cell(row=1, column=col_idx).value
                if h_val:
                    key = str(h_val).strip()
                    val, err = get_val_sdr_strict(source_data_pack, key, 0)
                    if val != "": ws[f"{col_letter}{curr}"] = val
            ws[f"F{curr}"] = i + 1

            val_refbp2, err_bp = get_val_sdr_strict(source_data_pack, "refBpDevice2", i, {'sheet': s_name, 'col': 'G'})
            if val_refbp2 is not None and str(val_refbp2).strip() != "":
                ws[f"G{curr}"] = f"1,1,{val_refbp2}"
            if err_bp:
                errors.append(err_bp)

            ws[f"H{curr}"] = i

            if mode == "TDD" and col_proto:
                ws.cell(row=curr, column=col_proto).value = "PHY LTE IR[1]"

            set_row_font(ws, curr, times_font)

    # --- 10. Common Sheets ---
    def set_ptp(ws, r, idx):
        val, err = get_val_sdr_strict(source_data_pack, "ptpDomain", idx, {'sheet': "IpClock", 'col': 'J'})
        if err:
            errors.append(err)
        else:
            ws[f"J{r}"] = val

    def set_clock(ws, r, idx):
        val_slot1, _ = get_val_sdr_strict(source_data_pack, "Slot1", 0)
        val_slot2, _ = get_val_sdr_strict(source_data_pack, "Slot2", 0)

        has_slot1 = val_slot1 is not None and str(val_slot1).strip() != ""
        has_slot2 = val_slot2 is not None and str(val_slot2).strip() != ""

        if has_slot1:
            ws[f"G{r}"] = 1
        elif has_slot2:
            ws[f"G{r}"] = 2

    common_sheets = [
        ("BandwidthResource", 1, None),
        ("IpClock", 1, set_ptp),
        ("Clock", 2, set_clock),
        ("OmcChannel", 1, None)
    ]
    for s_name, rows, action in common_sheets:
        ws = get_sheet_by_name_fuzzy(wb, s_name)
        if ws:
            prepare_template_rows(ws, 6, rows)
            for i in range(rows):
                curr = 6 + i
                fill_base_cols(ws, curr, s_name)
                if action: action(ws, curr, i)
                set_row_font(ws, curr, times_font)

    # --- 11. IrAntGroup (仅在 TDD 模式下操作) ---
    if mode == "TDD":
        s_name = "IrAntGroup"
        ws = get_sheet_by_name_fuzzy(wb, s_name)
        if ws:
            count = ru_rows
            prepare_template_rows(ws, 6, count)
            for i in range(count):
                curr = 6 + i
                fill_base_cols(ws, curr, s_name)
                ws[f"F{curr}"] = i + 1
                ws[f"I{curr}"] = i + 1

                ws[f"G{curr}"] = 1

                val_ru, err = get_val_sdr_strict(source_data_pack, "RUDevice", i, {'sheet': s_name, 'col': 'H'})
                if err: errors.append(err)
                if val_ru:
                    ws[f"H{curr}"] = val_ru

                for col_letter in ['J', 'K']:
                    col_idx = column_index_from_string(col_letter)
                    h_val = ws.cell(row=1, column=col_idx).value
                    if h_val:
                        key = str(h_val).strip()
                        val, err = get_val_sdr_strict(source_data_pack, key, i)
                        if val != "": ws[f"{col_letter}{curr}"] = val
                set_row_font(ws, curr, times_font)

    # 【全局标点净化】
    sanitize_punctuation(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, errors


# ==============================================================================
# 4. 核心处理逻辑 - CFG Radio (FDD & TDD)
# ==============================================================================

def process_cfg_radio_fdd(template_path, source_data_pack):
    if not os.path.exists(template_path):
        return None, [f"本地模版缺失: {template_path}"]

    wb = openpyxl.load_workbook(template_path)
    errors = []

    try:
        val, _ = get_val_sdr_strict(source_data_pack, "cellnum", 0)
        cell_num = int(val) if val else 1
    except:
        cell_num = 1

    # --- 1. ENBFunction ---
    s_name = "ENBFunction"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        r = 6
        count = 1
        prepare_template_rows(ws, 6, count)
        v_sub, err1 = get_val_sdr_strict(source_data_pack, "SubNetwork", 0, {'sheet': s_name, 'col': 'C'})
        v_me, err2 = get_val_sdr_strict(source_data_pack, "ManagedElement", 0, {'sheet': s_name, 'col': 'D'})
        v_ne, err3 = get_val_sdr_strict(source_data_pack, "NE_Name", 0, {'sheet': s_name, 'col': 'E'})
        if err1: errors.append(err1)
        if err2: errors.append(err2)
        if err3: errors.append(err3)

        ws[f"C{r}"] = v_sub
        ws[f"D{r}"] = v_me
        ws[f"E{r}"] = v_ne
        ws[f"H{r}"] = v_ne
        ws[f"I{r}"] = v_ne
        ws[f"J{r}"] = v_me
        ws[f"L{r}"] = v_me
        set_row_font(ws, r, times_font)

    # --- 2. Cell4GFDD ---
    s_name = "Cell4GFDD"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = cell_num
        prepare_template_rows(ws, 6, count)
        for i in range(count):
            curr = 6 + i

            v_sub, err1 = get_val_sdr_strict(source_data_pack, "SubNetwork", 0, {'sheet': s_name, 'col': 'C'})
            v_me, err2 = get_val_sdr_strict(source_data_pack, "ManagedElement", 0, {'sheet': s_name, 'col': 'D'})
            v_ne, err3 = get_val_sdr_strict(source_data_pack, "NE_Name", 0, {'sheet': s_name, 'col': 'E'})
            if err1: errors.append(err1)
            if err2: errors.append(err2)
            if err3: errors.append(err3)

            ws[f"C{curr}"] = v_sub
            ws[f"D{curr}"] = v_me
            ws[f"E{curr}"] = v_ne
            ws[f"F{curr}"] = v_me

            ranges = [
                range(21, 28),  # U-AA
                range(8, 10),  # H-I
                range(42, 49),  # AP-AV
                range(38, 40)  # AL, AM
            ]

            for rng in ranges:
                for col_idx in rng:
                    col_letter = get_column_letter(col_idx)
                    h_val = ws.cell(row=1, column=col_idx).value
                    if h_val:
                        key = str(h_val).strip()
                        val, err = get_val_sdr_strict(source_data_pack, key, i, {'sheet': s_name, 'col': col_letter})
                        if err:
                            errors.append(err)
                        elif val != "":
                            ws.cell(row=curr, column=col_idx).value = val

            fixed_map = {
                "L": "userLabel2", "Q": "pci", "S": "tac",
                "AF": "LONGITUDE", "AG": "LATITUDE",
                "AH": "rootSequenceIndex",
                "AK": "rfAppMode", "AN": "refBpDevice"
            }

            for col, key in fixed_map.items():
                val, err = get_val_sdr_strict(source_data_pack, key, i, {'sheet': s_name, 'col': col})
                if err:
                    errors.append(err)
                else:
                    ws[f"{col}{curr}"] = val

            ws[f"AI{curr}"] = ws[f"H{curr}"].value
            set_row_font(ws, curr, times_font)

    # 【全局标点净化】
    sanitize_punctuation(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, errors


def process_cfg_radio_tdd(template_path, source_data_pack, rancm_io=None):
    if not os.path.exists(template_path):
        return None, [f"本地模版缺失: {template_path}"]

    wb = openpyxl.load_workbook(template_path)
    errors = []

    try:
        val, _ = get_val_sdr_strict(source_data_pack, "cellnum", 0)
        cell_num = int(val) if val else 1
    except:
        cell_num = 1

    rru_grid = source_data_pack[2]
    rru_search_index = source_data_pack[5]

    generated_ref_sdr = []
    if rancm_io is not None:
        try:
            rancm_io.seek(0)
            rancm_wb = openpyxl.load_workbook(rancm_io, data_only=True)
            ir_ws = get_sheet_by_name_fuzzy(rancm_wb, "IrAntGroup")
            if ir_ws:
                ref_col = 8
                for c in range(1, ir_ws.max_column + 1):
                    h_val = str(ir_ws.cell(row=1, column=c).value or "").strip().lower()
                    if h_val == "refsdrdevicegroup":
                        ref_col = c
                        break
                for r in range(6, ir_ws.max_row + 1):
                    val = ir_ws.cell(row=r, column=ref_col).value
                    generated_ref_sdr.append(val if val is not None else "")
            rancm_io.seek(0)
        except Exception as e:
            errors.append(f"联动读取 RANCM 表 IrAntGroup 失败: {str(e)}")

    # --- 1. ENBFunction (TDD版) ---
    s_name = "ENBFunction"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        r = 6
        count = 1
        prepare_template_rows(ws, 6, count)
        v_sub, err1 = get_val_sdr_strict(source_data_pack, "SubNetwork", 0, {'sheet': s_name, 'col': 'C'})
        v_me, err2 = get_val_sdr_strict(source_data_pack, "ManagedElement", 0, {'sheet': s_name, 'col': 'D/H/L'})
        v_ne, err3 = get_val_sdr_strict(source_data_pack, "NE_Name", 0, {'sheet': s_name, 'col': 'E/I/J'})
        if err1: errors.append(err1)
        if err2: errors.append(err2)
        if err3: errors.append(err3)

        ws[f"C{r}"] = v_sub
        ws[f"D{r}"] = v_me
        ws[f"E{r}"] = v_ne

        ws[f"I{r}"] = v_ne
        ws[f"J{r}"] = v_ne

        ws[f"H{r}"] = v_me
        ws[f"L{r}"] = v_me

        set_row_font(ws, r, times_font)

    # --- 2. Cell4GTDD ---
    s_name = "Cell4GTDD"
    ws = get_sheet_by_name_fuzzy(wb, s_name)
    if ws:
        count = cell_num
        prepare_template_rows(ws, 6, count)
        for i in range(count):
            curr = 6 + i

            v_sub, err1 = get_val_sdr_strict(source_data_pack, "SubNetwork", 0, {'sheet': s_name, 'col': 'C'})
            v_me, err2 = get_val_sdr_strict(source_data_pack, "ManagedElement", 0, {'sheet': s_name, 'col': 'D/F'})
            v_ne, err3 = get_val_sdr_strict(source_data_pack, "NE_Name", 0, {'sheet': s_name, 'col': 'E'})
            if err1: errors.append(err1)
            if err2: errors.append(err2)
            if err3: errors.append(err3)

            ws[f"C{curr}"] = v_sub
            ws[f"D{curr}"] = v_me
            ws[f"E{curr}"] = v_ne
            ws[f"F{curr}"] = v_me

            val_cellLocalId, err = get_val_sdr_strict(source_data_pack, "cellLocalId", i,
                                                      {'sheet': s_name, 'col': 'H/I'})
            if val_cellLocalId != "":
                ws[f"H{curr}"] = val_cellLocalId
                ws[f"I{curr}"] = val_cellLocalId
            if err: errors.append(err)

            val_userLabel, err = get_val_sdr_strict(source_data_pack, "userLabel2", i, {'sheet': s_name, 'col': 'L'})
            if val_userLabel != "": ws[f"L{curr}"] = val_userLabel
            if err: errors.append(err)

            mappings = {
                "Q": "pci",
                "S": "tac",
                "U": "freqBandInd",
                "V": "earfcnUl",
                "W": "bandWidthDl",
                "Y": "bandWidthUl",
                "AE": "rootSequenceIndex",
                "AJ": "refBpDevice2",
                "AK": "cellMod",
                "AL": "cpSpeRefSigPwr",
                "AM": "upActAntBitmapSeq",
                "AN": "anttoPortMap",
                "AO": "isDelNbrAndRelation"
            }

            for col, key in mappings.items():
                val, err = get_val_sdr_strict(source_data_pack, key, i, {'sheet': s_name, 'col': col})
                if err:
                    errors.append(err)
                elif val != "":
                    ws[f"{col}{curr}"] = val

            ws[f"AF{curr}"] = i + 1

            if i < len(generated_ref_sdr) and str(generated_ref_sdr[i]).strip() != "":
                ws[f"AI{curr}"] = f"{generated_ref_sdr[i]}:1"
            else:
                errors.append(f"Sheet[{s_name}] 列[AI] 参数[refIrAntGroup]: 跨表联动失败，RANCM 对应有效数据不足。")
                ws[f"AI{curr}"] = f"{i + 1}:1"

            set_row_font(ws, curr, times_font)

    # 【全局标点净化】
    sanitize_punctuation(wb)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, errors


# ==============================================================================
# 5. 主界面
# ==============================================================================

st.title("📡 SDR 开站数据配置工具")
st.markdown("---")

if "processed" not in st.session_state:
    st.session_state.processed = False

sdr_file = st.file_uploader("📂 请上传 SDR 开站配置工单 (xlsx)", type=["xlsx"])

if sdr_file is not None and getattr(st.session_state, 'last_file', None) != sdr_file.name:
    st.session_state.processed = False
    st.session_state.last_file = sdr_file.name

if sdr_file:
    if not st.session_state.processed:
        if st.button("🚀 开始生成数据"):
            with st.spinner("数据引擎疯狂运算中，请稍候..."):
                try:
                    source_data_pack, headers, err = get_sdr_source_data(sdr_file)

                    if err:
                        st.error(f"❌ 读取源文件失败: {err}")
                    else:
                        mode_val, _ = get_val_sdr_strict(source_data_pack, "FDD/TDD", 0)
                        mode = "FDD"
                        if mode_val and "TDD" in str(mode_val).upper():
                            mode = "TDD"
                        elif mode_val and "FDD" in str(mode_val).upper():
                            mode = "FDD"

                        rancm_io, err_rancm = process_sdr_rancm(TEMPLATE_RANCM, source_data_pack, mode)

                        if mode == "FDD":
                            radio_name = TEMPLATE_FDD
                            radio_io, err_radio = process_cfg_radio_fdd(TEMPLATE_FDD, source_data_pack)
                        else:
                            radio_name = TEMPLATE_TDD
                            radio_io, err_radio = process_cfg_radio_tdd(TEMPLATE_TDD, source_data_pack, rancm_io)

                        all_errors = (err_rancm or []) + (err_radio or [])

                        if all_errors:
                            st.error("❌ 数据生成中止！检测到源文件数据或模版缺失，请按照提示检查：")
                            with st.expander("📋 点击查看详细错误列表", expanded=True):
                                for e in all_errors:
                                    st.write(f"🚫 {e}")
                        else:
                            st.session_state.rancm_io = rancm_io.getvalue()
                            st.session_state.radio_io = radio_io.getvalue()
                            st.session_state.mode = mode
                            st.session_state.radio_name = radio_name
                            st.session_state.headers = headers
                            st.session_state.processed = True
                            st.rerun()

                except Exception as e:
                    st.error("❌ 程序在处理数据时发生崩溃！请检查文件格式或联系开发者。")
                    st.exception(e)

    if st.session_state.processed:
        st.info(f"🔍 识别到制式模式: **{st.session_state.mode}**")
        st.success("✅ 数据生成成功！请点击下方按钮分别下载（可多次点击）。")

        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                label="📥 下载 RANCM 结果",
                data=st.session_state.rancm_io,
                file_name=f"Result_{TEMPLATE_RANCM}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        with col2:
            st.download_button(
                label=f"📥 下载 {st.session_state.mode} 配置",
                data=st.session_state.radio_io,
                file_name=f"Result_{st.session_state.radio_name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        st.markdown("---")
        if st.button("🔄 重新生成数据"):
            st.session_state.processed = False
            st.rerun()

        with st.expander("🔎 查看源文件表头识别"):
            st.write(st.session_state.headers)