import openpyxl
from openpyxl.utils import column_index_from_string
from copy import copy


def process_itbbu_template(file_path, output_path):
    print(f"正在加载文件: {file_path} ...")
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        print("错误：找不到文件，请确认路径。")
        return

    # === 配置：定义默认参数值字典 ===
    # 注意：需求中提到“从重选模版中按照第一行表头内容复制填充”
    # 由于我没有外部模版数据，这里定义一个字典。
    # 你需要在这里填入 H列 到 AY列 对应的【参数名: 标准值】
    # 如果找不到对应值，脚本目前会留空。
    DEFAULT_PARAM_VALUES = {
        "example_param_name": "0",
        "qRxLevMin": "-115",
        # 在此处添加更多具体的参数名和对应的默认值
    }

    # ==========================================
    # 任务 1: 操作第三个Sheet (CUEUtranCellFDDLTE)
    # ==========================================
    ws3 = wb.worksheets[2]  # 索引从0开始，2代表第三个
    print(f"正在处理 Sheet 3: {ws3.title}")

    # 锁定列索引 (openpyxl是从1开始)
    col_O = column_index_from_string('O')
    col_AW = column_index_from_string('AW')
    col_GK = column_index_from_string('GK')

    # 从第6行开始遍历
    for row in ws3.iter_rows(min_row=6, max_row=ws3.max_row):
        # 检查该行是否有内容（以第2列B列是否有值为例，或者检查全部）
        if all(c.value is None for c in row):
            continue

        # 修改指定列
        ws3.cell(row=row[0].row, column=col_O).value = "容量自适应[4]"
        ws3.cell(row=row[0].row, column=col_AW).value = "解关断[0]"
        ws3.cell(row=row[0].row, column=col_GK).value = "打开[1]"
        ws3.cell(row=row[0].row, column=1).value = "M"  # A列

    # ==========================================
    # 任务 2: 操作第五个Sheet (RatPriCnParaFDDLTE)
    # ==========================================
    ws5 = wb.worksheets[4]
    print(f"正在处理 Sheet 5: {ws5.title}")

    col_H = column_index_from_string('H')
    col_I = column_index_from_string('I')

    # 用于存储从Sheet 5提取的数据，供Sheet 6和8使用
    # 格式: list of [col_B_val, col_C_val, col_D_val, col_E_val, col_F_val]
    source_data_rows = []

    for row in ws5.iter_rows(min_row=6, max_row=ws5.max_row):
        # 简单判空，假设B列（索引1）必须有值
        if row[1].value is None:
            continue

        # 修改操作
        ws5.cell(row=row[0].row, column=col_H).value = "100"
        ws5.cell(row=row[0].row, column=col_I).value = "0"

        # 提取 B(1), C(2), D(3), E(4), F(5) 的值 (0-based index in row tuple)
        # 注意 row[i] 对应的是 cell 对象
        row_data = [row[1].value, row[2].value, row[3].value, row[4].value, row[5].value]
        source_data_rows.append(row_data)

    # ==========================================
    # 任务 3: 操作第六个Sheet (GeranMeasFDDLTE)
    # ==========================================
    ws6 = wb.worksheets[5]
    print(f"正在处理 Sheet 6: {ws6.title}")

    # 逻辑：通常这种操作是清空旧数据写入新数据，或者是追加。
    # 这里假设从第6行开始覆盖或写入
    start_row = 6

    # 获取表头映射 (假设第一行是参数名)
    # 我们需要填充 H(8) 到 AY(51)
    header_map_ws6 = {}
    for col in range(8, 52):  # 8 to 51 (H to AY)
        header_val = ws6.cell(row=1, column=col).value
        if header_val:
            header_map_ws6[col] = header_val.strip()

    for i, data in enumerate(source_data_rows):
        current_row = start_row + i

        # 1. A列填充 "A"
        ws6.cell(row=current_row, column=1).value = "A"

        # 2. 复制 B-F (对应 col 2-6)
        # data list index 0->B, 1->C, 2->D, 3->E, 4->F
        ws6.cell(row=current_row, column=2).value = data[0]
        ws6.cell(row=current_row, column=3).value = data[1]
        ws6.cell(row=current_row, column=4).value = data[2]
        ws6.cell(row=current_row, column=5).value = data[3]

        # 3. 处理 F列 (DN): 原内容 + 后缀
        original_dn = data[4] if data[4] else ""
        new_dn = f"{original_dn},GeranMeasFDDLTE=1"
        ws6.cell(row=current_row, column=6).value = new_dn

        # 4. 填充 H-AY (参数模板填充)
        for col_idx, param_name in header_map_ws6.items():
            # 尝试从默认值字典中获取，如果没有则留空或填默认值
            if param_name in DEFAULT_PARAM_VALUES:
                ws6.cell(row=current_row, column=col_idx).value = DEFAULT_PARAM_VALUES[param_name]
            else:
                # 如果没有配置对应值，可以选择填入 "TODO" 或者保持 None
                pass

                # ==========================================
    # 任务 4: 操作第七个Sheet (GsmReselectionFDDLTE)
    # ==========================================
    ws7 = wb.worksheets[6]
    print(f"正在处理 Sheet 7: {ws7.title}")

    col_J = column_index_from_string('J')
    col_K = column_index_from_string('K')
    col_L = column_index_from_string('L')

    for row in ws7.iter_rows(min_row=6, max_row=ws7.max_row):
        if row[0].value is None and row[1].value is None:  # 简单判空
            continue

        ws7.cell(row=row[0].row, column=col_J).value = "1[3]"
        ws7.cell(row=row[0].row, column=col_K).value = "1[3]"
        ws7.cell(row=row[0].row, column=col_L).value = 1  # 数值类型

    # ==========================================
    # 任务 5: 操作第八个Sheet (GsmRslParaFDD)
    # ==========================================
    ws8 = wb.worksheets[7]
    print(f"正在处理 Sheet 8: {ws8.title}")

    # 逻辑同 Sheet 6
    # 填充 H(8) 到 AX(50)
    header_map_ws8 = {}
    for col in range(8, 51):  # 8 to 50 (H to AX)
        header_val = ws8.cell(row=1, column=col).value
        if header_val:
            header_map_ws8[col] = header_val.strip()

    for i, data in enumerate(source_data_rows):
        current_row = start_row + i

        # A列
        ws8.cell(row=current_row, column=1).value = "A"

        # B-E列
        ws8.cell(row=current_row, column=2).value = data[0]
        ws8.cell(row=current_row, column=3).value = data[1]
        ws8.cell(row=current_row, column=4).value = data[2]
        ws8.cell(row=current_row, column=5).value = data[3]

        # F列 (注意：根据你的需求描述，这里也是加 ,GeranMeasFDDLTE=1)
        # 如果是笔误，通常这里应该是 ,GsmRslParaFDD=1。但我按你的要求写。
        original_dn = data[4] if data[4] else ""
        new_dn = f"{original_dn},GeranMeasFDDLTE=1"
        ws8.cell(row=current_row, column=6).value = new_dn

        # 填充 H-AX
        for col_idx, param_name in header_map_ws8.items():
            if param_name in DEFAULT_PARAM_VALUES:
                ws8.cell(row=current_row, column=col_idx).value = DEFAULT_PARAM_VALUES[param_name]

    print("保存文件...")
    wb.save(output_path)
    print(f"处理完成！文件已保存为: {output_path}")


# === 执行部分 ===
if __name__ == "__main__":
    # 在这里修改你的文件名
    input_file = "ITBBU重选模版.xlsx"
    output_file = "ITBBU重选模版_处理后.xlsx"

    # 检查当前目录下是否有该文件，防止报错
    import os

    if os.path.exists(input_file):
        process_itbbu_template(input_file, output_file)
    else:
        print(f"请将待处理的 '{input_file}' 放入当前文件夹，或修改代码中的文件路径。")